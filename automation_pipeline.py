# Fixed automation_pipeline.py
import pandas as pd
import numpy as np
import os
from datetime import datetime

RAW_PATH= 'data/superstore.csv'
CLEANED_PATH= 'data/cleaned_superstore.csv'
EXCEL_PATH='reports/task2_sql_results.xlsx'
LOG_PATH= 'scripts/reports/pipeline_run_log.txt'
RUN_TIMESTAMP=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
os.makedirs('scripts/data', exist_ok=True)
os.makedirs('scripts/reports', exist_ok=True)

def log(msg):
    line=f'[{RUN_TIMESTAMP}] {msg}'
    print(line)
    with open(LOG_PATH,'a',encoding='utf-8') as f:f.write(line+'\n')

def load_data(path=RAW_PATH):
    log(f'Loading {path}')
    return pd.read_csv(path)

def clean_data(df):
    df=df.copy().drop_duplicates()
    for c in ['Order Date','Ship Date']:
        if c in df.columns:
            df[c]=pd.to_datetime(df[c],errors='coerce')
    num=df.select_dtypes(include=[np.number]).columns
    obj=df.select_dtypes(include=['object','string']).columns
    for c in num:
        df[c]=df[c].fillna(df[c].median())
    for c in obj:
        m=df[c].mode(dropna=True)
        df[c]=df[c].fillna(m.iloc[0] if not m.empty else '')
    if 'Postal Code' in df.columns:
        df['Postal Code']=df['Postal Code'].fillna(0).astype(int).astype(str).str.zfill(5)
    for c in ['Sales','Profit']:
        if c in df.columns:
            q1,q3=df[c].quantile([0.25,0.75]);iqr=q3-q1
            df=df[(df[c]>=q1-3*iqr)&(df[c]<=q3+3*iqr)]
    if 'Order Date' in df.columns:
        df=df.dropna(subset=['Order Date'])
        df['Year']=df['Order Date'].dt.year
        df['Month']=df['Order Date'].dt.to_period('M').astype(str)
    if {'Profit','Sales'}.issubset(df.columns):
        df['Profit Margin %']=np.where(df['Sales']!=0,(df['Profit']/df['Sales']*100).round(2),0)
    return df

def save_cleaned(df): df.to_csv(CLEANED_PATH,index=False)

def calculate_kpis(df):
    ts=df['Sales'].sum()
    overall=pd.DataFrame([{
        'Run Timestamp':RUN_TIMESTAMP,
        'Total Sales ($)':round(ts,2),
        'Total Profit ($)':round(df['Profit'].sum(),2),
        'Total Orders':df['Order ID'].nunique(),
        'Total Customers':df['Customer ID'].nunique(),
        'Avg Order Value ($)':round(df.groupby('Order ID')['Sales'].sum().mean(),2),
        'Overall Profit Margin (%)':round(df['Profit'].sum()/ts*100,2) if ts else 0,
        'Avg Discount':round(df['Discount'].mean(),3)
    }])
    by_cat=df.groupby('Category').agg(Total_Sales=('Sales','sum'),Total_Profit=('Profit','sum'),Order_Count=('Order ID','nunique')).reset_index()
    by_cat['Profit Margin %']=np.where(by_cat['Total_Sales']!=0,(by_cat['Total_Profit']/by_cat['Total_Sales']*100).round(2),0)
    by_reg=df.groupby('Region').agg(Total_Sales=('Sales','sum'),Total_Profit=('Profit','sum'),Order_Count=('Order ID','nunique')).reset_index()
    monthly=df.groupby('Month').agg(Monthly_Sales=('Sales','sum'),Monthly_Profit=('Profit','sum')).reset_index().sort_values('Month')
    sub=df.groupby('Sub-Category').agg(Total_Sales=('Sales','sum'),Total_Profit=('Profit','sum')).reset_index().sort_values('Total_Profit')
    top=df.groupby(['Customer ID','Customer Name','Segment']).agg(Lifetime_Sales=('Sales','sum'),Lifetime_Profit=('Profit','sum'),Total_Orders=('Order ID','nunique')).reset_index().sort_values('Lifetime_Sales',ascending=False).head(10)
    return overall,by_cat,by_reg,monthly,sub,top

def export_to_excel(kpis):
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(EXCEL_PATH,engine='openpyxl') as w:
        names=['Overall KPIs','By Category','By Region','Monthly Trend','Sub-Category','Top 10 Customers']
        for df,name in zip(kpis,names):
            df.to_excel(w,sheet_name=name,index=False)
        for ws in w.sheets.values():
            for col in ws.iter_cols():
                width=min(max(len(str(c.value)) if c.value is not None else 0 for c in col)+4,40)
                ws.column_dimensions[get_column_letter(col[0].column)].width=width

def run_pipeline():
    try:
        df=load_data()
        df=clean_data(df)
        save_cleaned(df)
        export_to_excel(calculate_kpis(df))
        log('Pipeline completed successfully.')
    except Exception as e:
        log(f'ERROR: {e}')
        raise

if __name__=='__main__':
    run_pipeline()
